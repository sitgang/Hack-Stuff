# -*- coding: utf-8 -*-
import requests, re, os
from openpyxl import Workbook
import optparse

def getID(url):
    for match in re.finditer(r"deal_id=[0-9]{4,}", requests.get(url).text):return match.group().split('=')[1]

def crawl(shopID,pages):
    comments = []
    page = 1
    while page <= pages:
        print("[+] Page %d saved"%page)
        data = requests.get("https://www.nuomi.com/pcindex/main/comment", params = {"dealId" : shopID, "page" : page}).json()
        for item in data["data"]["list"]:
            comment = {}
            comment['name'] = item['nickname'].encode('utf8')
            comment['score'] = item['score']
            comment['create_time'] = item['create_time']
            comment['update_time'] = item['update_time']
            comment['content'] = item['content'].encode('utf8')
            comment['reply'] = ""
            if len(item['reply']) != 0:
                for reply in item['reply']:
                    comment['reply'] = reply['content'].encode('utf8')
                    break
            comments.append(comment)
        page += 1
    return comments

def save(comments,shopID):
    filename = os.getcwd() + os.sep + "NuomiShop%s.xlsx"%shopID
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = u"create_time"
    ws.cell(row=1, column=2).value = u"update_time"
    ws.cell(row=1, column=3).value = u"name"
    ws.cell(row=1, column=4).value = u"score"
    ws.cell(row=1, column=5).value = u"content"
    ws.cell(row=1, column=6).value = u"reply"
    
    for i in range(0, len(comments)):
        ws.cell(row=i+2, column=1).value = comments[i]['create_time']
        ws.cell(row=i+2, column=2).value = comments[i]['update_time']
        ws.cell(row=i+2, column=3).value = comments[i]['name']
        ws.cell(row=i+2, column=4).value = comments[i]['score']
        ws.cell(row=i+2, column=5).value = comments[i]['content']
        ws.cell(row=i+2, column=6).value = comments[i]['reply']
    
    if os.path.exists(filename):
        os.remove(filename)
    
    wb.save(filename)
    print("[:)] All Done!")
    print("[!] Saved to %s"%filename)
    

def main():
    parser = optparse.OptionParser('usage %prog -u'+\
        '<shop url> -p <pages>')
    parser.add_option('-u',dest='shopURL',type='string',\
        help='specify shop url')
    parser.add_option('-p',dest='pages',type='int',\
        help='specify pages to crawl')
    (options,args) = parser.parse_args()
    shopURL = options.shopURL
    pages = options.pages
    if (pages ==None) | (shopURL == None):
        print('[-] You must specify a shopURL and pages to crawl.')
        exit(0)
  
    shopID = getID(shopURL)
    comments = crawl(shopID,pages)
    save(comments,shopID)

if __name__ =="__main__":
    main()